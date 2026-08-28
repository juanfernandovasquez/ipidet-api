import io
import re
from markupsafe import Markup, escape
from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from starlette.types import ASGIApp, Scope, Receive, Send
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import webapp.db as pdb
import webapp.portal_db as portal_db
import webapp.portal_router as portal_routes
import webapp.auth as auth
import webapp.mailer as mailer
import webapp.scheduler as scheduler
from config.settings import SECRET_KEY

app = FastAPI(title="IPIDET Admin")

# CORS: solo permite llamadas desde el servidor WordPress (server-to-server)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://ipidet.org"],
    allow_methods=["GET", "POST"],
    allow_headers=["Authorization", "Content-Type", "X-WC-Webhook-Signature"],
)

_PUBLIC_PATHS = {"/login", "/logout"}
_PUBLIC_PREFIXES = ("/webhook/", "/api/portal/")

class _AuthMiddleware:
    """Middleware ASGI puro — compatible con SessionMiddleware sin interferir en la cookie."""
    def __init__(self, app: ASGIApp):
        self.app = app

    async def __call__(self, scope: Scope, receive: Receive, send: Send):
        if scope["type"] != "http":
            await self.app(scope, receive, send)
            return
        path = scope.get("path", "")
        if path in _PUBLIC_PATHS or any(path.startswith(p) for p in _PUBLIC_PREFIXES):
            await self.app(scope, receive, send)
            return
        session = scope.get("session", {})
        if not session.get("user_email"):
            response = RedirectResponse(f"/login?next={path}", status_code=302)
            await response(scope, receive, send)
            return
        role = session.get("user_role", "viewer")
        if role != "admin":
            section = auth.section_for_path(path)
            if section == "__admin__":
                response = RedirectResponse("/", status_code=302)
                await response(scope, receive, send)
                return
            if section and section not in session.get("user_permisos", []):
                response = RedirectResponse("/?sin_acceso=1", status_code=302)
                await response(scope, receive, send)
                return
        await self.app(scope, receive, send)

app.add_middleware(_AuthMiddleware)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, https_only=False, same_site="lax")
templates = Jinja2Templates(directory="webapp/templates")

_URL_RE = re.compile(r'(https?://[^\s\|\]>\"\']+)')

def _autolink(text: str) -> Markup:
    if not text:
        return Markup("")
    def _replace(m):
        url = m.group(1)
        short = url if len(url) <= 50 else url[:47] + "…"
        return f'<a href="{escape(url)}" target="_blank" rel="noopener" class="text-blue-600 hover:underline break-all">{escape(short)}</a>'
    return Markup(_URL_RE.sub(_replace, str(escape(text))))

templates.env.filters["autolink"] = _autolink

def _format_fecha(value) -> str:
    """Formatea fecha ISO string o datetime a DD/MM/YYYY."""
    if not value:
        return '—'
    try:
        s = str(value)[:10]  # "YYYY-MM-DD"
        y, m, d = s.split('-')
        return f"{d}/{m}/{y}"
    except Exception:
        return str(value)

templates.env.filters["format_fecha"] = _format_fecha

STATUS_LABELS = {
    "pagado":        ("Pagado",        "green"),
    "debe":          ("Debe",          "red"),
    "fraccionamiento":("Fraccionamiento","amber"),
    "exonerado":     ("Exonerado",     "gray"),
    "no_aplica":     ("N/A",           "gray"),
    "pendiente":     ("Pendiente",     "yellow"),
    "retirar":       ("Retirar",       "slate"),
    "en_revision":   ("En revisión",   "purple"),
    "revisar":       ("Revisar",       "orange"),
    "parcial":       ("Parcial",       "blue"),
}

def _ctx(request: Request, **kwargs):
    return {
        "status_labels": STATUS_LABELS,
        "medios_pago": pdb.MEDIOS_PAGO,
        "bancos": pdb.BANCOS,
        "current_user_email":   request.session.get("user_email", ""),
        "current_user_role":    request.session.get("user_role", ""),
        "current_user_permisos":request.session.get("user_permisos", []),
        "secciones": auth.SECCIONES,
        **kwargs,
    }


@app.on_event("startup")
async def _startup():
    pdb.seed_companies()
    auth.seed_admin()
    import asyncio
    asyncio.create_task(scheduler.run_scheduler())


# ── Dashboard ─────────────────────────────────────────────────────────────────

_SECTION_PATHS = [
    ("members", "/members"), ("billing", "/billing"),
    ("fraccionamientos", "/fraccionamientos"), ("finanzas", "/finanzas"),
    ("faqs", "/faqs"), ("marketing", "/marketing"),
]

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    role     = request.session.get("user_role", "viewer")
    permisos = request.session.get("user_permisos", [])
    if role != "admin" and "dashboard" not in permisos:
        for sec, path in _SECTION_PATHS:
            if sec in permisos:
                return RedirectResponse(path, status_code=302)
        return HTMLResponse("<h2 style='font-family:sans-serif;padding:2rem'>Sin acceso asignado. Contacta al administrador.</h2>")
    stats = pdb.get_stats()
    return templates.TemplateResponse(request, "dashboard.html", _ctx(request, stats=stats))


# ── Padrón ────────────────────────────────────────────────────────────────────

@app.get("/members", response_class=HTMLResponse)
async def members_list(
    request: Request,
    search: str = "",
    estado: str = "",
    pago: str = "",
    ubicacion: str = "",
    page: int = 1,
):
    docs, total = pdb.get_members(search, estado, pago, ubicacion, page)
    return templates.TemplateResponse(request, "members.html", _ctx(request,
        members=docs, total=total,
        search=search, estado=estado, pago=pago, ubicacion=ubicacion,
        page=page, per_page=50,
        total_pages=max(1, (total + 49) // 50),
    ))


@app.post("/members/nuevo")
async def member_nuevo(
    apellidos:       str = Form(...),
    nombres:         str = Form(...),
    titulo:          str = Form(""),
    email:           str = Form(""),
    celular:         str = Form(""),
    centro_trabajo:  str = Form(""),
    ubicacion:       str = Form(""),
    fecha_ingreso:   str = Form(""),
    notas:           str = Form(""),
):
    member_id = pdb.create_member(
        apellidos=apellidos, nombres=nombres, titulo=titulo,
        email=email, celular=celular, centro_trabajo=centro_trabajo,
        ubicacion=ubicacion, fecha_ingreso=fecha_ingreso, notas=notas,
    )
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.get("/members/{member_id}", response_class=HTMLResponse)
async def member_detail(request: Request, member_id: str):
    doc = pdb.get_member(member_id)
    if not doc:
        return RedirectResponse("/members")
    return templates.TemplateResponse(request, "member.html", _ctx(request, member=doc))


@app.post("/members/{member_id}/emails/toggle")
async def toggle_email(member_id: str, email: str = Form(...), estado: str = Form(...)):
    pdb.update_email_status(member_id, email, estado)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/emails/set-principal")
async def set_email_principal(member_id: str, email: str = Form(...)):
    pdb.set_email_principal(member_id, email)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/emails/add")
async def add_email(member_id: str, new_email: str = Form(...)):
    if new_email.strip():
        pdb.add_email(member_id, new_email.strip())
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/notes")
async def update_notes(member_id: str, notas: str = Form(...)):
    pdb.update_member_notes(member_id, notas)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/comentarios/add")
async def add_comentario(member_id: str, texto: str = Form(...)):
    if texto.strip():
        pdb.add_comentario(member_id, texto)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/comentarios/{comentario_id}/delete")
async def delete_comentario(member_id: str, comentario_id: str):
    pdb.delete_comentario(member_id, comentario_id)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


@app.post("/members/{member_id}/estado")
async def update_member_estado(member_id: str, estado: str = Form(...)):
    pdb.update_member_estado(member_id, estado)
    return RedirectResponse(f"/members/{member_id}", status_code=303)


# ── Cobranzas ─────────────────────────────────────────────────────────────────

@app.get("/billing/facturacion", response_class=HTMLResponse)
async def facturacion(
    request: Request,
    periodo: str = "2026",
    search: str = "",
):
    pendientes = pdb.get_comprobantes_pendientes(periodo, search)
    return templates.TemplateResponse(request, "facturacion.html", _ctx(request,
        pendientes=pendientes, periodo=periodo, search=search,
        total=len(pendientes),
    ))


@app.get("/billing/facturacion/empresa", response_class=HTMLResponse)
async def facturacion_empresa(request: Request, periodo: str = "2026"):
    return templates.TemplateResponse(request, "facturacion_empresa.html",
                                      _ctx(request, periodo=periodo))


@app.get("/api/facturacion/pendientes")
async def api_pendientes_socio(member_id: str, periodo: str = "2026"):
    return pdb.get_pendientes_socio(member_id, periodo)


@app.get("/api/facturacion/socio-info")
async def api_socio_info(member_id: str, periodo: str = "2026"):
    """Devuelve el estado actual del pago de un socio para el período indicado."""
    p = pdb.payments_col.find_one({"member_id": member_id, "periodo": periodo})
    if not p:
        return {"payment_id": None, "estado": "sin_registro", "empresa_pagadora": ""}
    return {
        "payment_id":      str(p["_id"]),
        "estado":          p.get("estado", ""),
        "empresa_pagadora": p.get("empresa_pagadora") or "",
    }


@app.post("/billing/facturacion/empresa/guardar")
async def guardar_factura_empresa(request: Request):
    data = await request.json()
    socios            = data.get("socios", [])   # [{member_id, payment_id}]
    num_comprobante   = data.get("num_comprobante", "").strip()
    tipo_comprobante  = data.get("tipo_comprobante", "")
    fecha_emision     = data.get("fecha_emision", "")
    monto             = float(data.get("monto") or 0)
    fecha_vencimiento = data.get("fecha_vencimiento", "")
    empresa           = data.get("empresa", "").strip()
    enviar_email      = data.get("enviar_email", False)
    periodo           = data.get("periodo", "2026")

    if not socios or not empresa or not num_comprobante or not tipo_comprobante or not fecha_emision:
        return JSONResponse({"error": "Faltan datos obligatorios"}, status_code=422)

    member_ids = []
    for s in socios:
        member_id  = s.get("member_id")
        payment_id = s.get("payment_id") or pdb.get_or_create_payment(member_id, periodo)
        pdb.mark_payment_empresa(payment_id, empresa, num_comprobante, tipo_comprobante, fecha_emision)
        member_ids.append(member_id)
        if enviar_email:
            pay = pdb.get_payment_with_member(payment_id)
            if pay and pay.get("email_principal"):
                html = mailer.tpl_comprobante(
                    nombre=pay["nombre_completo"],
                    num_comprobante=num_comprobante,
                    tipo=tipo_comprobante,
                    periodo=periodo,
                    fecha_emision=fecha_emision,
                )
                try:
                    await mailer.send_email(
                        to=pay["email_principal"],
                        subject=f"Comprobante IPIDET {periodo} — {num_comprobante}",
                        html_body=html,
                    )
                except Exception:
                    pass

    # Crear registro de crédito automáticamente
    if fecha_vencimiento and monto > 0:
        n = len(member_ids)
        concepto = f"Membresías {periodo} — {n} socio{'s' if n != 1 else ''}"
        pdb.create_factura_credito(
            empresa=empresa,
            numero_factura=num_comprobante,
            monto=monto,
            fecha_emision=fecha_emision,
            fecha_vencimiento=fecha_vencimiento,
            concepto=concepto,
            socios=member_ids,
        )

    return {"ok": True, "registrados": len(socios)}


@app.post("/billing/{payment_id}/emitir-comprobante")
async def emitir_comprobante(
    payment_id: str,
    tipo: str             = Form(...),
    numero: int | None    = Form(None),
    num_comprobante: str  = Form(...),
    tipo_comprobante: str = Form(""),
    fecha_emision: str    = Form(""),
    enviar_email: str     = Form(""),
    periodo: str          = Form("2026"),
    search: str           = Form(""),
):
    num = num_comprobante.strip()
    pdb.emitir_comprobante(payment_id, tipo, numero, num, tipo_comprobante, fecha_emision)
    if enviar_email == "on":
        pay = pdb.get_payment_with_member(payment_id)
        if pay and pay.get("email_principal"):
            html = mailer.tpl_comprobante(
                nombre=pay["nombre_completo"],
                num_comprobante=num,
                tipo=tipo_comprobante,
                periodo=pay.get("periodo", periodo),
                fecha_emision=fecha_emision,
            )
            try:
                await mailer.send_email(
                    to=pay["email_principal"],
                    subject=f"Comprobante IPIDET {periodo} — {num}",
                    html_body=html,
                )
            except Exception:
                pass  # no bloquear el flujo si el email falla
    return RedirectResponse(f"/billing/facturacion?periodo={periodo}&search={search}", status_code=303)

@app.get("/billing", response_class=HTMLResponse)
async def billing(
    request: Request,
    periodo: str = "2026",
    estado: str = "",
    empresa: str = "",
    search: str = "",
    comprobante_emitido: str = "",
    page: int = 1,
):
    docs, total = pdb.get_payments(periodo, estado, empresa, search, page, comprobante_emitido=comprobante_emitido)
    return templates.TemplateResponse(request, "billing.html", _ctx(request,
        payments=docs, total=total,
        periodo=periodo, estado=estado, empresa=empresa, search=search,
        comprobante_emitido=comprobante_emitido,
        page=page, per_page=50,
        total_pages=max(1, (total + 49) // 50),
    ))


@app.get("/billing/export")
async def billing_export(
    periodo: str = "2026",
    estado: str = "",
    empresa: str = "",
    search: str = "",
):
    docs = pdb.get_payments_export(periodo, estado, empresa, search)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cobranzas {periodo}"

    headers = ["ID Socio", "Apellidos", "Nombres", "Centro de Trabajo", "Email",
               "Período", "Estado", "Empresa", "Pagado por", "Fecha Pago", "Medio Pago", "Cuotas"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    STATUS_ES = {
        "pagado": "Pagado", "debe": "Debe", "fraccionamiento": "Fraccionamiento",
        "exonerado": "Exonerado", "no_aplica": "N/A", "pendiente": "Pendiente",
        "retirar": "Retirar", "en_revision": "En revisión", "revisar": "Revisar",
    }
    for row, d in enumerate(docs, 2):
        cuotas = d.get("cuotas", [])
        cuotas_pagadas = sum(1 for c in cuotas if c.get("estado") == "pagado")
        cuotas_str = f"{cuotas_pagadas}/{len(cuotas)} pagadas" if cuotas else ""
        ws.cell(row=row, column=1,  value=d.get("member_id", ""))
        ws.cell(row=row, column=2,  value=d.get("apellidos", ""))
        ws.cell(row=row, column=3,  value=d.get("nombres", ""))
        ws.cell(row=row, column=4,  value=d.get("centro_trabajo", ""))
        ws.cell(row=row, column=5,  value=d.get("email_principal", ""))
        ws.cell(row=row, column=6,  value=d.get("periodo", ""))
        ws.cell(row=row, column=7,  value=STATUS_ES.get(d.get("estado", ""), d.get("estado", "")))
        ws.cell(row=row, column=8,  value=d.get("empresa_pagadora") or "")
        ws.cell(row=row, column=9,  value=d.get("pagado_por") or "")
        ws.cell(row=row, column=10, value=d.get("fecha_pago") or "")
        ws.cell(row=row, column=11, value=d.get("medio_pago") or "")
        ws.cell(row=row, column=12, value=cuotas_str)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"cobranzas_{periodo}"
    if estado:
        filename += f"_{estado}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/billing/{payment_id}/update")
async def update_payment(
    payment_id: str,
    estado: str = Form(...),
    empresa: str = Form(""),
    fecha_pago: str = Form(""),
    medio: str = Form(""),
    pagado_por: str = Form(""),
    num_comprobante: str = Form(""),
    tipo_comprobante: str = Form(""),
    fecha_emision_comprobante: str = Form(""),
    link_constancia: str = Form(""),
    banco_origen: str = Form(""),
    comprobante_emitido: str = Form(""),
    redirect_to: str = Form("/billing"),
):
    comp_emit = True if comprobante_emitido == "true" else (False if comprobante_emitido == "false" else None)
    pdb.update_payment(payment_id, estado, empresa or None, fecha_pago or None,
                       medio or None, pagado_por or None, num_comprobante or None,
                       tipo_comprobante or None, link_constancia or None,
                       banco_origen or None, comp_emit,
                       fecha_emision_comprobante or None)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/cuotas/objetivo")
async def set_cuota_objetivo(
    payment_id: str,
    monto_objetivo: float = Form(...),
    redirect_to: str = Form("/billing"),
):
    pdb.set_monto_objetivo(payment_id, monto_objetivo)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/cuotas/add")
async def add_cuota(
    payment_id: str,
    monto: float = Form(...),
    fecha_venc: str = Form(""),
    redirect_to: str = Form("/billing"),
):
    pdb.add_cuota(payment_id, monto, fecha_venc or None)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/cuotas/{numero}/update")
async def update_cuota(
    payment_id: str,
    numero: int,
    estado: str = Form(...),
    fecha_pago: str = Form(""),
    medio: str = Form(""),
    num_comprobante: str = Form(""),
    tipo_comprobante: str = Form(""),
    link_constancia: str = Form(""),
    banco_origen: str = Form(""),
    monto: str = Form(""),
    fecha_venc: str = Form(""),
    redirect_to: str = Form("/billing"),
):
    monto_f = float(monto) if monto.strip() else None
    pdb.update_cuota(payment_id, numero, estado, fecha_pago or None,
                     medio or None, num_comprobante or None,
                     tipo_comprobante or None, link_constancia or None,
                     banco_origen or None, monto_f, fecha_venc or None)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/cuotas/{numero}/delete")
async def delete_cuota(
    payment_id: str,
    numero: int,
    redirect_to: str = Form("/billing"),
):
    pdb.delete_cuota(payment_id, numero)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/parciales/init")
async def init_parcial(
    payment_id: str,
    monto_total: float = Form(...),
    redirect_to: str = Form("/billing"),
):
    pdb.set_monto_total(payment_id, monto_total)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/parciales/add")
async def add_pago_parcial(
    payment_id: str,
    monto: float = Form(...),
    fecha_pago: str = Form(""),
    medio: str = Form(""),
    num_comprobante: str = Form(""),
    tipo_comprobante: str = Form(""),
    link_constancia: str = Form(""),
    banco_origen: str = Form(""),
    redirect_to: str = Form("/billing"),
):
    pdb.add_pago_parcial(payment_id, monto, fecha_pago or None, medio or None,
                         num_comprobante or None, tipo_comprobante or None,
                         link_constancia or None, banco_origen or None)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/parciales/{numero}/update")
async def update_pago_parcial(
    payment_id: str,
    numero: int,
    monto: float = Form(...),
    fecha_pago: str = Form(""),
    medio: str = Form(""),
    num_comprobante: str = Form(""),
    tipo_comprobante: str = Form(""),
    link_constancia: str = Form(""),
    banco_origen: str = Form(""),
    redirect_to: str = Form("/billing"),
):
    pdb.update_pago_parcial(payment_id, numero, monto, fecha_pago or None,
                            medio or None, num_comprobante or None,
                            tipo_comprobante or None, link_constancia or None,
                            banco_origen or None)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/parciales/{numero}/delete")
async def delete_pago_parcial(
    payment_id: str,
    numero: int,
    redirect_to: str = Form("/billing"),
):
    pdb.delete_pago_parcial(payment_id, numero)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/billing/{payment_id}/socio/estado")
async def update_socio_estado_from_billing(
    payment_id: str,
    estado_socio: str = Form(...),
    redirect_to: str = Form("/billing"),
):
    from bson import ObjectId
    p = pdb.payments_col.find_one({"_id": ObjectId(payment_id)}, {"member_id": 1})
    if p and p.get("member_id"):
        pdb.update_member_estado(p["member_id"], estado_socio)
    return RedirectResponse(redirect_to, status_code=303)


# ── Marketing ─────────────────────────────────────────────────────────────────

@app.get("/marketing", response_class=HTMLResponse)
async def marketing(
    request: Request,
    titulo: str = "",
    ubicacion: str = "",
    estado_pago: str = "",
    empresa: str = "",
    periodo: str = "2026",
):
    stats = pdb.get_marketing_stats(periodo)
    emails = pdb.get_marketing_emails(titulo, ubicacion, estado_pago, empresa, periodo)
    titulos = sorted(set(t for t in pdb.members_col.distinct("titulo") if t))
    ubicaciones = sorted(set(u for u in pdb.members_col.distinct("ubicacion") if u))
    return templates.TemplateResponse(request, "marketing.html", _ctx(request,
        stats=stats, emails=emails,
        titulos=titulos, ubicaciones=ubicaciones,
        titulo=titulo, ubicacion=ubicacion,
        estado_pago=estado_pago, empresa=empresa, periodo=periodo,
    ))


@app.get("/marketing/export")
async def marketing_export(
    titulo: str = "",
    ubicacion: str = "",
    estado_pago: str = "",
    empresa: str = "",
    periodo: str = "2026",
):
    emails = pdb.get_marketing_emails(titulo, ubicacion, estado_pago, empresa, periodo)
    buf = io.StringIO()
    buf.write("Nombre,Email Principal,Email Secundario,Título,Ubicación,Centro de Trabajo,Estado Pago\n")
    for r in emails:
        nombre   = r["nombre"].replace('"', '""')
        email    = r["email"].replace('"', '""')
        email2   = r.get("email_secundario", "").replace('"', '""')
        tit      = r["titulo"].replace('"', '""')
        ubic     = r["ubicacion"].replace('"', '""')
        ct       = r["centro_trabajo"].replace('"', '""')
        ep       = r["estado_pago"].replace('"', '""')
        buf.write(f'"{nombre}","{email}","{email2}","{tit}","{ubic}","{ct}","{ep}"\n')
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=marketing_{periodo}.csv"},
    )


# ── Finanzas (mockup) ─────────────────────────────────────────────────────────

@app.get("/finanzas", response_class=HTMLResponse)
async def finanzas_mockup(request: Request):
    return templates.TemplateResponse(request, "finanzas_mockup.html", _ctx(request))


# ── Fraccionamientos ─────────────────────────────────────────────────────────

@app.get("/fraccionamientos", response_class=HTMLResponse)
async def fraccionamientos(
    request: Request,
    periodo: str = "2026",
    alerta: str = "",
    search: str = "",
    page: int = 1,
):
    from datetime import date as _date_cls
    docs, total, stats = pdb.get_fraccionamientos(periodo, alerta, search, page)
    return templates.TemplateResponse(request, "fraccionamientos.html", _ctx(request,
        fraccionamientos=docs, total=total, stats=stats,
        periodo=periodo, alerta=alerta, search=search,
        page=page, per_page=50,
        total_pages=max(1, (total + 49) // 50),
        today=_date_cls.today().isoformat(),
    ))


# ── FAQs ──────────────────────────────────────────────────────────────────────

@app.get("/faqs", response_class=HTMLResponse)
async def faqs_list(request: Request, search: str = "", category: str = ""):
    docs = pdb.get_faqs(search, category)
    categories = pdb.get_faq_categories()
    return templates.TemplateResponse(request, "faqs.html", _ctx(request,
        faqs=docs, categories=categories,
        search=search, category=category,
    ))


@app.post("/faqs/add")
async def add_faq(
    question: str = Form(...),
    answer: str = Form(...),
    category: str = Form(...),
):
    pdb.save_faq(question, answer, category)
    return RedirectResponse("/faqs", status_code=303)


@app.post("/faqs/{faq_id}/delete")
async def delete_faq(faq_id: str):
    pdb.delete_faq(faq_id)
    return RedirectResponse("/faqs", status_code=303)


# ── Empresas (página de gestión) ─────────────────────────────────────────────

@app.get("/empresas", response_class=HTMLResponse)
async def empresas_page(request: Request, search: str = "", tipo: str = ""):
    empresas = pdb.get_all_companies(search, tipo)
    return templates.TemplateResponse(request, "empresas.html", _ctx(request,
        empresas=empresas, search=search, tipo=tipo,
    ))


@app.post("/empresas/add")
async def empresas_add(
    request: Request,
    nombre:            str = Form(...),
    ruc:               str = Form(""),
    razon_social:      str = Form(""),
    tipo:              str = Form("empresa"),
    contacto_nombre:   str = Form(""),
    contacto_email:    str = Form(""),
    contacto_telefono: str = Form(""),
):
    pdb.add_company(nombre, ruc, razon_social, tipo,
                    contacto_nombre, contacto_email, contacto_telefono)
    return RedirectResponse("/empresas", status_code=303)


@app.post("/empresas/{company_id}/update")
async def empresas_update(
    company_id: str,
    nombre:            str = Form(...),
    ruc:               str = Form(""),
    razon_social:      str = Form(""),
    tipo:              str = Form("empresa"),
    contacto_nombre:   str = Form(""),
    contacto_email:    str = Form(""),
    contacto_telefono: str = Form(""),
):
    pdb.update_company(company_id, nombre, ruc, razon_social, tipo,
                       contacto_nombre, contacto_email, contacto_telefono)
    return RedirectResponse("/empresas", status_code=303)


@app.post("/empresas/{company_id}/delete")
async def empresas_delete(company_id: str):
    pdb.delete_company(company_id)
    return RedirectResponse("/empresas", status_code=303)


# ── API empresas (autocomplete) ───────────────────────────────────────────────

@app.get("/api/companies")
async def api_companies(q: str = ""):
    return pdb.get_companies(q)


@app.post("/api/companies")
async def api_add_company(nombre: str = Form(...), ruc: str = Form(""), tipo: str = Form("empresa")):
    pdb.add_company(nombre, ruc, tipo=tipo)
    return {"ok": True}


@app.delete("/api/companies/{company_id}")
async def api_delete_company(company_id: str):
    pdb.delete_company(company_id)
    return {"ok": True}


# ── API (para uso del bot) ────────────────────────────────────────────────────

@app.get("/api/members")
async def api_members(search: str = "", estado: str = "", pago: str = ""):
    docs, total = pdb.get_members(search, estado, pago)
    return {"total": total, "members": docs}


@app.get("/api/members/{member_id}")
async def api_member(member_id: str):
    doc = pdb.get_member(member_id)
    return doc or JSONResponse({"error": "not found"}, status_code=404)


@app.get("/api/payments")
async def api_payments(periodo: str = "2026", estado: str = ""):
    docs, total = pdb.get_payments(periodo, estado)
    return {"total": total, "payments": docs}


# ── Email / Comunicaciones ────────────────────────────────────────────────────

@app.post("/members/{member_id}/send-email")
async def send_member_email(member_id: str, request: Request):
    data    = await request.json()
    asunto  = (data.get("asunto") or "").strip()
    cuerpo  = (data.get("cuerpo") or "").strip()
    if not asunto or not cuerpo:
        return JSONResponse({"ok": False, "error": "Asunto y mensaje son obligatorios."}, status_code=422)
    member = pdb.get_member(member_id)
    if not member:
        return JSONResponse({"ok": False, "error": "Socio no encontrado."}, status_code=404)
    email_to = next(
        (e["email"] for e in member.get("emails", [])
         if e.get("estado") == "habilitado" and e.get("principal")),
        next((e["email"] for e in member.get("emails", [])
              if e.get("estado") == "habilitado"), None),
    )
    if not email_to:
        return JSONResponse({"ok": False, "error": "El socio no tiene email habilitado."}, status_code=422)
    cuerpo_html = cuerpo.replace("\n", "<br>")
    nombre = f"{member.get('nombres','')} {member.get('apellidos','')}".strip()
    html = mailer._base_html(f"""
      <p style="color:#475569;line-height:1.7;white-space:pre-line">{cuerpo_html}</p>
    """)
    try:
        await mailer.send_email(to=email_to, subject=asunto, html_body=html)
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.post("/api/email/test")
async def email_test(to: str = Form(...)):
    """Envía un correo de prueba para verificar la conexión Brevo."""
    try:
        await mailer.send_email(
            to=to,
            subject="Prueba de conexión IPIDET – Brevo",
            html_body=mailer.tpl_bienvenida("Equipo IPIDET"),
        )
        return {"ok": True, "to": to}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ── Facturas a crédito ────────────────────────────────────────────────────────

@app.get("/billing/credito", response_class=HTMLResponse)
async def billing_credito(request: Request, empresa: str = "", estado: str = ""):
    # Resolver RUC → nombre de empresa si el filtro parece un RUC
    empresa_filtro = empresa
    if empresa.strip():
        nombres = pdb._resolve_empresa_nombres(empresa)
        if nombres:
            empresa_filtro = nombres[0]
    facturas = pdb.get_facturas_credito(empresa_filtro, estado)
    stats    = pdb.get_credito_stats()
    empresas = pdb.get_all_companies()
    return templates.TemplateResponse(request, "credito.html", _ctx(request,
        facturas=facturas, stats=stats, empresas=empresas,
        empresa=empresa, estado=estado,
    ))

@app.post("/billing/credito/nueva")
async def credito_nueva(
    empresa: str         = Form(...),
    numero_factura: str  = Form(...),
    monto: float         = Form(...),
    fecha_emision: str   = Form(...),
    fecha_vencimiento: str = Form(...),
    concepto: str        = Form(""),
):
    pdb.create_factura_credito(empresa, numero_factura, monto,
                                fecha_emision, fecha_vencimiento, concepto)
    return RedirectResponse("/billing/credito", status_code=303)

@app.post("/billing/credito/{factura_id}/estado")
async def credito_estado(factura_id: str, estado: str = Form(...),
                          fecha_cobro: str = Form("")):
    pdb.update_factura_credito_estado(factura_id, estado, fecha_cobro)
    return RedirectResponse("/billing/credito", status_code=303)

@app.post("/billing/credito/{factura_id}/delete")
async def credito_delete(factura_id: str):
    pdb.delete_factura_credito(factura_id)
    return RedirectResponse("/billing/credito", status_code=303)


@app.post("/billing/credito/{factura_id}/comentarios/add")
async def credito_add_comentario(factura_id: str, request: Request):
    data = await request.json()
    texto = (data.get("texto") or "").strip()
    if not texto:
        return JSONResponse({"error": "Texto vacío"}, status_code=422)
    comentario = pdb.add_comentario_credito(factura_id, texto)
    return {"ok": True, "comentario": comentario}


@app.post("/billing/credito/{factura_id}/comentarios/{idx}/delete")
async def credito_delete_comentario(factura_id: str, idx: int):
    pdb.delete_comentario_credito(factura_id, idx)
    return {"ok": True}


# ── Comunicaciones ────────────────────────────────────────────────────────────

@app.get("/comunicaciones", response_class=HTMLResponse)
async def comunicaciones(request: Request):
    history  = pdb.get_comunicaciones_history()
    titulos  = pdb.get_member_titulos()
    ubicaciones = pdb.get_member_ubicaciones()
    empresas = pdb.get_companies()
    return templates.TemplateResponse(request, "comunicaciones.html", _ctx(request,
        history=history, titulos=titulos, ubicaciones=ubicaciones, empresas=empresas,
    ))


@app.post("/api/comunicaciones/preview")
async def comunicaciones_preview(request: Request):
    data        = await request.json()
    periodo     = data.get("periodo", "2026")
    estados     = data.get("estados_pago", [])
    empresa     = data.get("empresa", "")
    ubicacion   = data.get("ubicacion", "")
    titulo      = data.get("titulo", "")
    destinatarios = pdb.get_comunicacion_destinatarios(
        periodo=periodo, estados_pago=estados or None,
        empresa=empresa, ubicacion=ubicacion, titulo=titulo,
    )
    return {"total": len(destinatarios), "destinatarios": destinatarios}


@app.post("/api/comunicaciones/enviar")
async def comunicaciones_enviar(request: Request):
    data          = await request.json()
    asunto        = (data.get("asunto") or "").strip()
    cuerpo        = (data.get("cuerpo") or "").strip()
    plantilla     = data.get("plantilla", "libre")
    destinatarios = data.get("destinatarios", [])
    filtros       = data.get("filtros", {})

    if not asunto or not cuerpo:
        return JSONResponse({"error": "Asunto y mensaje son obligatorios."}, status_code=422)
    if not destinatarios:
        return JSONResponse({"error": "No hay destinatarios seleccionados."}, status_code=422)

    cuerpo_html = cuerpo.replace("\n", "<br>")
    html_body = mailer._base_html(f'<p style="color:#475569;line-height:1.7">{cuerpo_html}</p>')

    mensajes = [{"to": d["email"], "subject": asunto, "html_body": html_body}
                for d in destinatarios if d.get("email")]

    enviados, fallidos, errores = await mailer.send_bulk(mensajes)

    if enviados == 0 and fallidos > 0:
        return JSONResponse(
            {"error": f"No se pudo enviar ningún correo. Error: {errores[0] if errores else 'desconocido'}"},
            status_code=500,
        )

    usuario = request.session.get("user_email", "")
    pdb.save_comunicacion_log(asunto, plantilla, filtros, destinatarios, usuario)

    return {"ok": True, "enviados": enviados, "fallidos": fallidos,
            "errores": errores if errores else []}


# ── Eventos ───────────────────────────────────────────────────────────────────

MESES_ES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}


@app.get("/eventos", response_class=HTMLResponse)
async def eventos_list(
    request: Request,
    search: str = "",
    estado: str = "",
    page: int = 1,
):
    eventos, total = pdb.get_eventos(search=search, estado=estado, page=page)
    # Agrupar por mes
    grupos: dict[str, list] = {}
    for ev in eventos:
        fecha = ev.get("fecha", "")
        clave = fecha[:7] if fecha else "Sin fecha"   # "YYYY-MM"
        if clave not in grupos:
            grupos[clave] = []
        grupos[clave].append(ev)
    # Enriquecer clave con nombre de mes
    grupos_list = []
    for clave, evs in grupos.items():
        if len(clave) == 7:
            anio, mes = clave.split("-")
            label = f"{MESES_ES.get(mes, mes)} {anio}"
        else:
            label = clave
        grupos_list.append({"clave": clave, "label": label, "eventos": evs})
    pages = (total + 39) // 40
    return templates.TemplateResponse(request, "eventos.html", _ctx(request,
        grupos=grupos_list, total=total, page=page, pages=pages,
        search=search, estado=estado,
    ))


@app.post("/eventos/nuevo")
async def evento_nuevo(
    request: Request,
    titulo:      str = Form(...),
    descripcion: str = Form(""),
    fecha:       str = Form(...),
    hora:        str = Form(""),
    lugar:       str = Form(""),
    cupo_max:    str = Form(""),
):
    cupo = int(cupo_max) if cupo_max.strip() else None
    pdb.create_evento(titulo, descripcion, fecha, hora, lugar, cupo)
    return RedirectResponse("/eventos", status_code=303)


@app.get("/eventos/{evento_id}", response_class=HTMLResponse)
async def evento_detalle(request: Request, evento_id: str):
    evento = pdb.get_evento(evento_id)
    if not evento:
        return RedirectResponse("/eventos", status_code=302)
    stats = pdb.get_evento_stats(evento_id)
    return templates.TemplateResponse(request, "evento_detalle.html", _ctx(request,
        evento=evento, stats=stats,
    ))


@app.post("/eventos/{evento_id}/update")
async def evento_update(
    request: Request,
    evento_id:   str,
    titulo:      str = Form(...),
    descripcion: str = Form(""),
    fecha:       str = Form(...),
    hora:        str = Form(""),
    lugar:       str = Form(""),
    cupo_max:    str = Form(""),
    estado:      str = Form("activo"),
):
    cupo = int(cupo_max) if cupo_max.strip() else None
    pdb.update_evento(evento_id, titulo, descripcion, fecha, hora, lugar, cupo, estado)
    return RedirectResponse(f"/eventos/{evento_id}", status_code=303)


@app.post("/eventos/{evento_id}/delete")
async def evento_delete(request: Request, evento_id: str):
    pdb.delete_evento(evento_id)
    return RedirectResponse("/eventos", status_code=303)


@app.post("/eventos/{evento_id}/inscribir")
async def evento_inscribir(
    request: Request,
    evento_id: str,
    member_id: str = Form(...),
):
    error = pdb.inscribir_socio(evento_id, member_id)
    if error:
        return JSONResponse({"error": error}, status_code=422)
    return JSONResponse({"ok": True})


@app.post("/eventos/{evento_id}/desinscribir/{member_id}")
async def evento_desinscribir(request: Request, evento_id: str, member_id: str):
    pdb.desinscribir_socio(evento_id, member_id)
    return RedirectResponse(f"/eventos/{evento_id}", status_code=303)


@app.post("/eventos/{evento_id}/asistencia/{member_id}")
async def evento_asistencia(
    request: Request,
    evento_id: str,
    member_id: str,
    asistio: str = Form(...),
):
    valor: bool | None = None if asistio == "" else (asistio == "true")
    pdb.marcar_asistencia(evento_id, member_id, valor)
    return RedirectResponse(f"/eventos/{evento_id}", status_code=303)


@app.get("/api/eventos")
async def api_eventos(limit: int = 5):
    """API pública para bots: próximos eventos activos."""
    eventos = pdb.get_eventos_proximos(limit=limit)
    return {"eventos": eventos}


@app.get("/api/eventos/{evento_id}")
async def api_evento_detalle(evento_id: str):
    """Detalle de evento con stats para bots."""
    stats = pdb.get_evento_stats(evento_id)
    if not stats:
        return JSONResponse({"error": "No encontrado"}, status_code=404)
    return stats


# ── Pendientes y Atención ─────────────────────────────────────────────────────

@app.get("/pendientes", response_class=HTMLResponse)
async def pendientes_list(
    request: Request,
    estado: str = "",
    prioridad: str = "",
    search: str = "",
):
    items = pdb.get_pendientes(estado, prioridad, search)
    stats = pdb.get_pendientes_stats()
    return templates.TemplateResponse(request, "pendientes.html", _ctx(request,
        items=items, stats=stats,
        estado=estado, prioridad=prioridad, search=search,
    ))


@app.post("/pendientes/add")
async def pendiente_add(
    titulo: str = Form(...),
    descripcion: str = Form(""),
    prioridad: str = Form("media"),
    member_id: str = Form(""),
    nombre_miembro: str = Form(""),
):
    if member_id.strip() and not nombre_miembro.strip():
        m = pdb.get_member(member_id.strip())
        if m:
            nombre_miembro = f"{m.get('apellidos','')} {m.get('nombres','')}".strip()
    pdb.create_pendiente(titulo, descripcion, prioridad, member_id, nombre_miembro)
    return RedirectResponse("/pendientes", status_code=303)


@app.post("/pendientes/{pendiente_id}/update")
async def pendiente_update(
    pendiente_id: str,
    titulo: str = Form(...),
    descripcion: str = Form(""),
    prioridad: str = Form("media"),
    estado: str = Form("pendiente"),
    member_id: str = Form(""),
    nombre_miembro: str = Form(""),
    redirect_to: str = Form("/pendientes"),
):
    if member_id.strip() and not nombre_miembro.strip():
        m = pdb.get_member(member_id.strip())
        if m:
            nombre_miembro = f"{m.get('apellidos','')} {m.get('nombres','')}".strip()
    pdb.update_pendiente(pendiente_id, titulo, descripcion, prioridad, estado,
                         member_id, nombre_miembro)
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/pendientes/{pendiente_id}/estado")
async def pendiente_estado(
    pendiente_id: str,
    estado: str = Form(...),
    redirect_to: str = Form("/pendientes"),
):
    from bson import ObjectId
    p = pdb.pendientes_col.find_one({"_id": ObjectId(pendiente_id)})
    if p:
        pdb.update_pendiente(
            pendiente_id,
            p.get("titulo", ""), p.get("descripcion", ""),
            p.get("prioridad", "media"), estado,
            p.get("member_id") or "", p.get("nombre_miembro") or "",
        )
    return RedirectResponse(redirect_to, status_code=303)


@app.post("/pendientes/{pendiente_id}/delete")
async def pendiente_delete(
    pendiente_id: str,
    redirect_to: str = Form("/pendientes"),
):
    pdb.delete_pendiente(pendiente_id)
    return RedirectResponse(redirect_to, status_code=303)


# ── Portal de socios ──────────────────────────────────────────────────────────
@app.get("/api/portal/member-status")
async def portal_member_status(
    request: Request,
    email: str = Query(...),
):
    from config.settings import PORTAL_SECRET
    auth_header = request.headers.get("authorization", "")
    if PORTAL_SECRET and auth_header != f"Bearer {PORTAL_SECRET}":
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    if not email or "@" not in email:
        return JSONResponse({"error": "Email inválido"}, status_code=400)
    return portal_routes.build_member_status(email)


@app.post("/webhook/woocommerce/order")
async def portal_wc_webhook(request: Request):
    return await portal_routes.handle_wc_webhook(request)


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, next: str = "/", error: str = ""):
    if request.session.get("user_email"):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse(request, "login.html", {"next": next, "error": error})


@app.post("/login")
async def login_post(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    next: str = Form("/"),
):
    user = auth.get_user(email)
    if not user or not auth.verify_password(password, user["password_hash"]):
        return templates.TemplateResponse(request, "login.html", {
            "next": next,
            "error": "Correo o contraseña incorrectos.",
        })
    request.session["user_email"]   = user["email"]
    request.session["user_role"]    = user["role"]
    request.session["user_permisos"]= user.get("permisos", auth.ALL_SECTIONS)
    return RedirectResponse(next if next.startswith("/") else "/", status_code=302)


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)


# ── Admin: gestión de usuarios ────────────────────────────────────────────────

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request):
    if request.session.get("user_role") != "admin":
        return RedirectResponse("/", status_code=302)
    users = auth.list_users()
    return templates.TemplateResponse(request, "users.html", _ctx(request, users=users))


@app.post("/admin/users/add")
async def admin_add_user(request: Request):
    if request.session.get("user_role") != "admin":
        return RedirectResponse("/", status_code=302)
    form = await request.form()
    email    = form.get("email", "")
    password = form.get("password", "")
    role     = form.get("role", "viewer")
    permisos = form.getlist("permisos")
    if not permisos and role == "viewer":
        permisos = auth.ALL_SECTIONS
    auth.create_user(email, password, role, permisos)
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/permisos")
async def admin_update_permisos(request: Request, user_id: str):
    if request.session.get("user_role") != "admin":
        return RedirectResponse("/", status_code=302)
    form = await request.form()
    permisos = form.getlist("permisos")
    auth.update_user_permisos(user_id, permisos)
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/delete")
async def admin_delete_user(request: Request, user_id: str):
    if request.session.get("user_role") != "admin":
        return RedirectResponse("/", status_code=302)
    auth.delete_user(user_id)
    return RedirectResponse("/admin/users", status_code=303)
